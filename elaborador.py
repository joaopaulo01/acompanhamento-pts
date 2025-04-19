from tkinter import filedialog
import webbrowser
import pandas as pd
import re, os
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException
from webdriver_manager.firefox import GeckoDriverManager
from xml.sax.saxutils import escape
import time
from tkcalendar import Calendar
from datetime import date
from datetime import datetime
from dateutil.relativedelta import relativedelta
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from openpyxl import load_workbook
import openpyxl

# Variáveis globais para armazenar os valores
usuario = ""
senha = ""
data = ""
caminho_arquivo = ""
recomendacoes_emitente = """- NENHUMA ATIVIDADE FORA DO ESCOPO DA PT/AR PODERÁ SER REALIZADA.

- EM CASO DE EMERGÊNCIAS: 8800 

- RAMAL SALA DE CONTROLE: 751-5856 

EM CASO DE ALARME SONORO CONTÍNUO POR MAIS DE 40 SEGUNDOS, PARALISE AS ATIVIDADES E SIGA AS ROTAS DE FUGA ATÉ O PONTO DE ENCONTRO LOCALIZADO ATRÁS DO ESTACIONAMENTO DO PRÉDIO DA OPERAÇÃO. 

- UTILIZAR ACESSO ADEQUADO À EXECUÇÃO DO TRABALHO; 

- MANTER DISTÂNCIA DAS PARTES MÓVEIS, ENERGIZADAS, AQUECIDAS OU PRESSURIZADAS DO EQUIPAMENTO; 

- MANTER ORDEM E LIMPEZA NO LOCAL; 

- PARAR O TRABALHO CASO MUDEM AS CONDIÇÕES AMBIENTAIS E/OU OPERACIONAIS."""

#FUNÇÃO PARA SELECIONAR ARQUIVO
def selecionar_arquivo():
    global caminho_arquivo
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
    )
    if caminho_arquivo:
        print(f"Arquivo selecionado: {caminho_arquivo}")

#FUNÇÃO PARA LER PLANILHA
def ler_planilha_excel(arquivo_excel):
    """
    Lê os dados da planilha Excel e armazena os valores em variáveis com base nos nomes das colunas.
    """
    try:
        # Carrega a planilha em um DataFrame
        df = pd.read_excel(arquivo_excel)
        colunas_dados = {}
        # Itera pelas colunas da planilha
        for coluna in df.columns:
            colunas_dados[coluna] = df[coluna].tolist()  # Armazena os valores da coluna como lista
        return colunas_dados
    except Exception as e:
        print(f"Erro ao ler a planilha Excel: {e}")
        return None

#FUNÇÃO PARA REALIZAR LOGIN
def realizar_login(driver, usuario, senha):
    """
    Função para realizar o login no sistema.
    """
    print("Iniciando processo de login...")
    
    # URL do site
    url = "https://portalspt.petrobras.com.br/portspt/pages/comum/cookieMenu.do"

    while True:  # Loop para repetir a solicitação de credenciais em caso de erro
        try:
            # Acessa o site
            driver.get(url)

            # Aguarda um tempo para garantir que o site carregue completamente
            time.sleep(3)

            # Alterna para o iframe onde está o formulário de login
            iframe = driver.find_element(By.CSS_SELECTOR, 'iframe[src^="/portspt/pages/public/noticias/consultar.do"]')
            driver.switch_to.frame(iframe)

            # Localiza os campos de login e senha
            username_field = driver.find_element(By.CSS_SELECTOR, 'form#LoginForm fieldset label input[name="j_username"]')
            password_field = driver.find_element(By.CSS_SELECTOR, 'form#LoginForm fieldset label input[name="j_password"]')

            # Preenche os campos com o usuário e a senha
            username_field.send_keys(usuario)  # Preenche o campo de usuário
            password_field.send_keys(senha)  # Preenche o campo de senha

            # Pressiona Enter para enviar o formulário
            password_field.send_keys(Keys.RETURN)

            # Aguarda um tempo para o site processar o login
            time.sleep(5)

            # Verifica se os campos de login ainda estão presentes
            try:
                username_field = driver.find_element(By.CSS_SELECTOR, 'form#LoginForm fieldset label input[name="j_username"]')
                password_field = driver.find_element(By.CSS_SELECTOR, 'form#LoginForm fieldset label input[name="j_password"]')
                print("Usuário ou senha incorretos. Por favor, tente novamente.")
                return  # Volta ao início do loop para solicitar novas credenciais
            except:
                # Se os campos de login não forem encontrados, o login foi bem-sucedido
                print("Login realizado com sucesso!")
                break  # Sai do loop

        except Exception as e:
            print(f"Erro ao realizar o login: {e}")
            continue  # Volta ao início do loop para tentar novamente

    # Volta ao contexto principal (fora do iframe de login)
    driver.switch_to.default_content()
    return driver

#FUNÇÃO PARA ELABORAR PERMISSÕES DE TRABALHO
def elaborar_pts(usuario, senha, data, caminho_arquivo, empresa_selecionada, area_selecionada, unidade_selecionada):
    print("Iniciando elaboração de PTs...")
    print(f"Usuário: {usuario}")
    print(f"Data: {data}")
    print(f"Caminho do arquivo: {caminho_arquivo}")
    print(f"Empresa: {empresa_selecionada}")
    print(f"Área: {area_selecionada}")
    print(f"Unidade: {unidade_selecionada}")

    if caminho_arquivo:
        dados_planilha = ler_planilha_excel(caminho_arquivo)
    else:
        print("Nenhum arquivo Excel foi selecionado.")
        return
    #"C:\Users\DMYN\Downloads\edgedriver_win64\msedgedriver.exe"
    #driver_path = "C:/Users/DMYN/Downloads/edgedriver_win64/msedgedriver.exe"
    #service = Service(driver_path)    
    #driver = None

    try:
        #driver = webdriver.Edge(service=service) <- EDGE
        #driver = webdriver.Firefox(service=Service(GeckoDriverManager().install())) #<- FIREFOX
        driver = webdriver.Firefox()
        driver = realizar_login(driver, usuario, senha)
        acessar_spt(driver, empresa_selecionada, area_selecionada, unidade_selecionada)
        clicar_permissao_trabalho(driver)
        elaborar_permissoes_trabalho(driver, dados_planilha, data, caminho_arquivo)
        salvar_informacoes_planilha(dados_planilha, caminho_arquivo)
        print("Elaboração finalizada com sucesso!")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
    """finally:
        if driver:
            driver.quit()"""

#FUNÇÃO PARA ACESSAR SPT
def acessar_spt(driver, empresa_selecionada, area_selecionada, unidade_selecionada):
    iframe_link = driver.find_element(By.CSS_SELECTOR, 'iframe[src^="/portspt/pages/public/noticias/consultar.do"]')
    driver.switch_to.frame(iframe_link)
    link_spt = driver.find_element(By.CSS_SELECTOR, 'a#linkSpt > img')
    link_spt.click()
    print("Clicou no linkSpt com sucesso.")
    driver.switch_to.default_content()
    iframe_main = driver.find_element(By.CSS_SELECTOR, 'iframe#main')
    driver.switch_to.frame(iframe_main)
    seletor_empresa = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "empresas"))
    )
    dropwdown = Select(seletor_empresa)
    dropwdown.select_by_visible_text(empresa_selecionada)
    time.sleep(2)
    seletor_area = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "areaNegocio"))
    )
    dropwdown = Select(seletor_area)
    dropwdown.select_by_visible_text(area_selecionada)
    time.sleep(2)
    seletor_unidades = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "unidades"))
    )
    dropwdown = Select(seletor_unidades)
    dropwdown.select_by_visible_text(unidade_selecionada)
    time.sleep(2)
    botao = driver.find_element(By.CSS_SELECTOR, 'input#sel_un\\:botao')
    botao.click()
    print("Botão 'sel_un:botao' pressionado com sucesso.")

#FUNÇÃO PARA CLICAR EM PERMISSÃO PARA TRABALHO
def clicar_permissao_trabalho(driver):
    permissao_trabalho = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//ul[@id='treemenu1']/li/strong[text()='Permissão para Trabalho']"))
    )
    permissao_trabalho.click()
    print("Clicou no elemento 'Permissão para Trabalho' com sucesso.")

#FUNÇÃO PARA ELABORAR PERMISSÕES DE TRABALHO
def elaborar_permissoes_trabalho(driver, dados_planilha, data, caminho_arquivo):
    for indice, acao in enumerate(dados_planilha["Ação"]):
        print(f"indice atual: {indice}")
        try:
            print(f"Ação: {acao}")
            if "Elaborar" in str(acao):
                elaborar_permissao = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "(//a[@title='Elaborar'])[2]"))
                )
                elaborar_permissao.click()
                
                elaborar_pt(driver, dados_planilha, indice, data, caminho_arquivo)
            else:
                print(f"PT Emitida")
        except Exception as e:
            print("Ordem já foi emitida alguma vez")

#FUNÇÃO PARA ELABORAR PT
def elaborar_pt(driver, dados_planilha, index, data, caminho_arquivo):
    time.sleep(2)
    campo_codigo_om = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//input[@id='codigoOm']"))
    )
    campo_codigo_om.clear()
    campo_codigo_om.send_keys(dados_planilha["Ordem"][index])
    time.sleep(2)
    botao_pesquisar = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@id='j_id271']"))
    )
    driver.execute_script("arguments[0].scrollIntoView(true);", botao_pesquisar)
    botao_pesquisar.click()
    tratar_pt_encerrada(driver, dados_planilha, index)
    selecionar_operacoes(driver, dados_planilha, index, data)
    #preencher_campos_pt(driver, dados_planilha, index, data)
    #salvar_informacoes_planilha(dados_planilha, caminho_arquivo)
    #salvar_pt(driver, dados_planilha, index, caminho_arquivo)

#FUNÇÃO PARA IDENTIFICAR SE A ORDEM ESTÁ ENCERRADA/NÃO PLANEJADA
def tratar_pt_encerrada(driver, dados_planilha, index):
    try:
        pt_emitida_salva = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='modalMensagemDiv']/center/input[@value='Fechar']"))
        )
        pt_emitida_salva.click()
        print("Ordem não disponível para Emissão de PT")
        return
    except Exception as e:
        print("PT disponível para elaboração. Iniciando...")

#FUNÇÃO PARA SELECIONAR OPERAÇÕES
def selecionar_operacoes(driver, dados_planilha, index, data):
    imagem_link = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//table[starts-with(@id, 'item_0020')]/tbody/tr/td/img"))
    )
    imagem_link.click()
    pt_aberta = None
    
    #VERIFICA O ESTADO DA PT (ABERTA, EMITIDA OU ENCERRADA)
    try:
        OP = dados_planilha["Operações"][index]
        if pd.notna(OP):
            OP = [v.strip() for v in re.split(r"[\n,e]", str(OP)) if v.strip()]
            print(f"Linha {index + 1}: {OP}")
        for valor in OP:
            valor_escape = escape(valor.strip())
            print(f"Linha {valor}: {valor_escape}")
            op_atual = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//span[contains(@id, 'j_id172:table_tabelaResultadoOMs:0:j_id261')][contains(text(), '0{valor_escape}')]"))
            )
            op_atual.click()
            print(f"Checkbox para {valor_escape} foi selecionado.")
            
        elaboracao = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@id='j_id172:elaborar' and @type='submit' and @value='Elaborar']"))
        )
        elaboracao.click()
        try:
            pt_elaborar = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "/html/body/center/div/div[2]/div[2]/form/fieldset/legend/span[2]"))
            )
            print("PT disponível para elaboração. Iniciando...")
            if "Elaborar PT" in pt_elaborar.text:
                print("Elaborando PT...") 
                preencher_campos_pt(driver, dados_planilha, index, data)
        except Exception as e:
            pt_ja_emitida = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[@id='modalMensagemDiv']/center/input[@value='Fechar']"))
            )
            pt_ja_emitida.click()
            print("Ordem não disponível para Emissão de PT")
        
        
        """elemento = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, 
                f"//*[contains(@id, 'j_id172:table_tabelaResultadoOMs:0:j_id261')]"))
        )
        
        # Verifica o estado do elemento encontrado
        if "Aberta" in elemento.text:
            print("PT Salva")
            dados_planilha["Ação"][index] = "Emitir"
            dados_planilha["PT Emitida"][index] = "PT Salva"
            pt_aberta = True
            return
        elif "Emitida" in elemento.get_attribute("title"):
            print("PT Emitida")
            dados_planilha["Ação"][index] = "Reemitir"
            dados_planilha["PT Emitida"][index] = "PT Emitida"
            pt_aberta = True
            return
        elif "Encerrada" in elemento.get_attribute("title"):
            print("PT Encerrada")
            dados_planilha["Ação"][index] = "Reemitir"
            dados_planilha["PT Emitida"][index] = "PT Encerrada"
            pt_aberta = True
            return
        else:
            # Caso nenhum dos estados seja encontrado
            print("PT Disponível para Elaboração")
            OP = dados_planilha["Operações"][index]
            if pd.notna(OP):
                OP = [v.strip() for v in re.split(r"[\n,e]", str(OP)) if v.strip()]
                print(f"Linha {index + 1}: {OP}")
            for valor in OP:
                valor_escape = escape(valor.strip())
                print(f"Linha {valor}: {valor_escape}")
                op_atual = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//span[contains(@id, 'j_id172:table_tabelaResultadoOMs:0:j_id261')][contains(text(), '0{valor_escape}')]"))
                )
                op_atual.click()
                print(f"Checkbox para {valor_escape} foi selecionado.")
                
            elaboracao = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@id='j_id172:elaborar' and @type='submit' and @value='Elaborar']"))
            )
            elaboracao.click()
            preencher_campos_pt(driver, dados_planilha, index, data)"""  
            
    except Exception as e:
        # Caso nenhum dos estados seja encontrado
        print("Algo errado na abertura da ordem {e}")

#FUNÇÃO PARA PREENCHER CAMPOS DA PT
def preencher_campos_pt(driver, dados_planilha, index, data):
    Gerencia = dados_planilha["Gerência Emitente"][index]
    if Gerencia == 'G&E/UN-TERM/UTE-NPI/OP':
        GerenciaEmitente = "G&E/UN-TERM/UTE-NPI/OP"
    else:
        GerenciaEmitente = "G&E/UN-TERM/UTE-NPI/MAN"
    try:
        seletor_gerencia = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "gerenciaCombo"))
        )
        dropwdown = Select(seletor_gerencia)
        dropwdown.select_by_visible_text(GerenciaEmitente)
    except Exception as e:
        print(f"Erro ao selecionar uma opção: {e}")

    campo_descricao = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//textarea[@id='campoTextdescricao']"))
    )
    campo_descricao.send_keys(dados_planilha["Descrição da PT"][index])

    PBS = dados_planilha["PBS"][index]
    if pd.notna(PBS):
        PBS = [v.strip() for v in re.split(r"[.\n,e]", str(PBS)) if v.strip()]
        print(f"Linha {index + 1}: {PBS}")
    try:
        for valorPBS in PBS:
            valor_PBS = escape(valorPBS.strip())
            valor_PBS = int(valor_PBS) - 6
            pbs_atual = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, f"//input[@id='checkbox_verificacaoBasicaSeguranca[{valor_PBS}]']"))
            )
            if not pbs_atual.is_selected():
                pbs_atual.click()
                print(f"Checkbox para {valor_PBS + 6} foi selecionado.")
            else:
                print(f"Checkbox {valor_PBS + 6} já estava marcada.")
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
        for valorPBS in PBS:
            valor_PBS = escape(valorPBS.strip())
            valor_PBS = int(valor_PBS) - 6
            if valor_PBS == 2 or valor_PBS == 3:
                selecionar_todos_pbs = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//input[contains(@onclick, 'PBS0{valor_PBS + 6}') and contains(@class, 'todos_PBS0{valor_PBS + 6}')]"))
                )
            else:
                selecionar_todos_pbs = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//input[contains(@onclick, 'PBS{valor_PBS + 6}') and contains(@class, 'todos_PBS{valor_PBS + 6}')]"))
                )
            if not selecionar_todos_pbs.is_selected():
                selecionar_todos_pbs.click()
                print(f"Checkbox para {valor_PBS + 6} foi selecionado.")
            else:
                print(f"Checkbox {valor_PBS + 6} já estava marcada.")
    except Exception as e:
        print(f"Não entrou no for pois {e}")
        
    campo_recomendacoes = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//textarea[@id='campoTextdescricao2']"))
    )
    campo_recomendacoes.clear()
    campo_recomendacoes.send_keys(recomendacoes_emitente)
    
    pt_ptt = "PT" if dados_planilha["Tipo\nPT/PTT"][index] == "PT" else "PTT"
    if pt_ptt == "PT":
        radio_pt = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='form_pt:tipo_pt']/tbody/tr/td[1]/label/input"))
        )
        radio_pt.click()
        print("Botão de 'PT' selecionado com sucesso.")
    else:
        radio_ptt = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='form_pt:tipo_pt']/tbody/tr/td[2]/label/input"))
        )
        radio_ptt.click()
        print("Botão de 'PTT' selecionado com sucesso.")
        
        data_inicio_ptt = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='form_pt:dataInicioInputDate']"))
        )
        data_inicio_ptt.clear()
        data_inicio_ptt.send_keys(data)
        
        data_fim = datetime.strptime(data, "%d/%m/%Y") + relativedelta(months=1)
        data_fim = data_fim.strftime("%d/%m/%Y")
        data_termino_ptt = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='form_pt:dataTerminoInputDate']"))
        )
        data_termino_ptt.clear()
        data_termino_ptt.send_keys(data_fim)

        horario_inicio = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='horaInicioValidade']"))
        )
        horario_inicio.clear()
        horario_inicio.send_keys("07:30")
        
        horario_fim = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='horaTerminoValidade']"))
        )
        horario_fim.clear()
        horario_fim.send_keys("17:00")
        
        
    
    CorretivaPreventiva = dados_planilha["Tipo de\nIntervenção"][index]
    try:
        seletor_tipo = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "form_pt:comboTipoIntervencao"))
        )
        dropwdown1 = Select(seletor_tipo)
        dropwdown1.select_by_visible_text(CorretivaPreventiva)
    except Exception as e:
        print(f"Erro ao selecionar uma opção: {e}")

    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
    )

    CriticidadePT = dados_planilha["Criticidade\nPMIC"][index]
    CriticidadePT = CriticidadePT.capitalize()
    print(CriticidadePT)
    try:
        seletor_tipo = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "form_pt:comboCriticidade"))
        )
        dropwdown2 = Select(seletor_tipo)
        dropwdown2.select_by_visible_text(CriticidadePT)
    except Exception as e:
        print(f"Erro ao selecionar uma opção: {e}")

    WebDriverWait(driver, 10).until(
        EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
    )

    QuenteFrio = dados_planilha["Forma de\nTrabalho"][index]
    try:
        seletor_tipo = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "comboFormaTrabalho"))
        )
        dropwdown3 = Select(seletor_tipo)
        dropwdown3.select_by_visible_text(QuenteFrio)
    except Exception as e:
        print(f"Erro ao selecionar uma opção: {e}")
        
    #//*[@id="checkEspacoConfinado"] <- Espaço confinado
    try:
        espacoConfinado = dados_planilha["Espaço\nConfinado"][index]
        if espacoConfinado.lower() == "sim":
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            ec_sim = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='checkEspacoConfinado']"))
            )
            ec_sim.click()
            print("Botão de 'Sim' de espaço confinado selecionado com sucesso.")
    except Exception as e:
        print(f"Erro ao selecionar uma opção: {e}")

    try:
        AlertaVermelho = dados_planilha["Alerta\nVermelho"][index]
        if AlertaVermelho.lower() == "sim":
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            radio_sim = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='radioAlertaVermelho' and @value='S']"))
            )
            radio_sim.click()
            print("Botão de 'Sim' de alerta vermelho selecionado com sucesso.")
        else:
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            radio_nao = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@name='radioAlertaVermelho' and @value='N']"))
            )
            radio_nao.click()
            print("Botão de 'Não' de alerta vermelho selecionado com sucesso.")
    except Exception as e:
        print(f"Erro ao selecionar o botão de rádio: {e}")
        
    #//*[@id="checkboxItemVerificacao_[1]"] <- Especialidades do Serviço (checkbox de ELETRICIDADE)  
    try:
        if dados_planilha["Especialidade"][index].lower() == "elétrica":
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            verify_list = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='checkboxItemVerificacao_[1]']"))
            )
            verify_list.click()
    except Exception as e:
        print(f"Sem especialidade a selecionar {e}")
    if pt_ptt == "PTT":
        ptc_tipo = "Representante da Empreiteira"
    else:
        ptc_tipo = "Requisitante**"
        
    try:
        requisitantes = dados_planilha["Requisitantes"][index]
        print(requisitantes)
        if pd.notna(requisitantes):
            requisitantes = [v.strip() for v in re.split(r"[\n,e]", str(requisitantes)) if v.strip()]
        for valor in requisitantes:
            valor_escape = escape(valor.strip())
            campo_requisitantes = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@id='form_pt:A']"))
            )
            campo_requisitantes.send_keys(valor_escape)
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            tipo_participante = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@id='comboTipoParticipacao']"))
            )
            dropdown4 = Select(tipo_participante)
            dropdown4.select_by_visible_text(ptc_tipo)
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            botao_incluir = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[contains(@class, 'botao_incluir') and contains(@onclick, 'abrirPopUpParticipante')]"))
            )
            botao_incluir.click()
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            try:
                #//*[@id="form_pt:tabelaParticipantesAr:0:j_id1047"] <- para tentar clicar no requisitante
                find_req = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, f"//*[@id='form_pt:tabelaParticipantesAr:0:j_id1047']"))
                )
                find_req.click()
                print(f"Requisidante {valor} adicionado")
            except Exception as e:
                problem_req = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//div[@id='modalMensagemDiv']/center/input[@value='Fechar']"))
                )
                problem_req.click()
                print(f"Problema com o requisitante {valor}")

        campo_data = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='form_pt:dtIniProgformInputDate']"))
        )
        campo_data.clear()
        campo_data.send_keys(data)

        campo_horario = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='hrIniProgform']"))
        )
        campo_horario.clear()
        campo_horario.send_keys("07:30")

        campo_linhas_executantes = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='inputLinhaCienciaEmitirform']"))
        )
        campo_linhas_executantes.click()
        campo_linhas_executantes.clear()
        campo_linhas_executantes.send_keys("5")

        botao_confirmar = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "(//input[@value='Confirmar' and @title='Confirmar' and @type='button'])[3]"))
        )
        botao_confirmar.click()

        botao_salvar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[@type='button' and @class='Button_green' and @value='Salvar' and @title='Salvar']"))
        )
        botao_salvar.click()

        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
        )

        tratar_operacao_fora_ar(driver, dados_planilha, index)
        capturar_numero_pt(driver, dados_planilha, index, caminho_arquivo)
    except Exception as e:
        print(f"Erro ao selecionar o botão de rádio: {e}")

#FUNÇÃO PARA TRATAR OPERAÇÃO FORA DA AR
def tratar_operacao_fora_ar(driver, dados_planilha, index):
    try:
        botao_operacao_fora_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='form_pt:cfmConfirmarPtComARSemRecomendacaoDiv']/center/input"))
        )
        operacao_fora_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='form_pt:cfmConfirmarPtComARSemRecomendacaoDiv']//span[2]"))
        )
        botao_operacao_fora_ar.click()
        print(f"Operação {operacao_fora_ar} está fora da AR")
        OperacaoForaAR = operacao_fora_ar.text
        dados_planilha["Observação"][index] = f"Operação {OperacaoForaAR} está fora da AR"
        return
    
    except Exception as e:
        print("AR está ok. PT Salva")

#FUNÇÃO PARA CAPTURAR NÚMERO DA PT
def capturar_numero_pt(driver, dados_planilha, index, caminho_arquivo):
    try:
        texto_pt_salva = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//table[@id='mensagens']/tbody/tr/td/span[starts-with(text(), 'Permissão para Trabalho nº')]"))
        )
        texto_completo = texto_pt_salva.text
        print(f"Texto completo capturado: {texto_completo}")
        match = re.search(r"Permissão para Trabalho nº (\S+\s\S+)", texto_completo)
        if match:
            valor_extraido = match.group(1)
            dados_planilha["PT Emitida"][index] = valor_extraido
            dados_planilha["Ação"][index] = "Emitir"
            print(f"O valor da PT que está na planilha é {dados_planilha['PT Emitida'][index]}")
            
        else:
            print("Não foi possível extrair o valor do texto.")
    except Exception as e:
        print(f"Erro ao capturar o texto: {e}")

#FUNÇÃO PARA SALVAR INFORMAÇÕES NA PLANILHA
def salvar_informacoes_planilha(dados_planilha, caminho_arquivo):
    try:
        #SALVAR PANDAS COMO EXCEL
        dados = pd.DataFrame(dados_planilha)
        print(dados)

        # Cria um objeto ExcelWriter
        with pd.ExcelWriter('SemanaAtual.xlsx', engine='openpyxl') as writer:
            dados.to_excel(writer, index=False, sheet_name='Sheet1')

            # Acessa o workbook e a worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']

            # Adiciona filtros na primeira linha
            for col in range(len(dados.columns)):
                worksheet.auto_filter.ref = worksheet.dimensions
            
            # Ajusta a largura das colunas automaticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter  # Obtém a letra da coluna
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Define a quebra de linha automática para todas as células
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

            # Aplica formatação condicional
            green_fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    if cell.column_letter == 'G' and cell.value != 'Não':
                        cell.fill = green_fill
                    if cell.column_letter == 'B' and cell.value == 'Emitir':
                        cell.fill = green_fill
    except Exception as e:
        print(f"Erro ao salvar informações na planilha: {e}")

#FUNÇÃO PARA CONSULTAR AR
def consultar_ar(usuario, senha, caminho_arquivo, empresa_selecionada, area_selecionada, unidade_selecionada):
    if caminho_arquivo:
        dados_planilha = ler_planilha_excel(caminho_arquivo)
    else:
        print("Nenhum arquivo Excel foi selecionado.")
        return

    # Caminho para o WebDriver do Edge
    #driver_path = "C:/Users/DMYN/Downloads/edgedriver_win64/msedgedriver.exe"
    #service = Service(driver_path)
    driver = None

    #INICIO DAS CONSULTAS DE AR
    try:
        #Inicializa o driver
        #driver = webdriver.Edge(service=service)
        driver = webdriver.Firefox()
        #driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()))

        # Realiza o login
        driver = realizar_login(driver, usuario, senha)

        # Após o login bem-sucedido, realiza as ações solicitadas
        try:
            # Alterna para o iframe onde está o link "linkSpt"
            iframe_link = driver.find_element(By.CSS_SELECTOR, 'iframe[src^="/portspt/pages/public/noticias/consultar.do"]')
            driver.switch_to.frame(iframe_link)

            # Clica no link "SPT"
            link_spt = driver.find_element(By.CSS_SELECTOR, 'a#linkSpt > img')
            link_spt.click()
            print("Clicou no linkSpt com sucesso.")

            # Alterna para o próximo iframe onde está o botão "sel_un:botao"
            driver.switch_to.default_content()  # Sai do iframe atual
            iframe_main = driver.find_element(By.CSS_SELECTOR, 'iframe#main')
            driver.switch_to.frame(iframe_main)
            
            seletor_empresa = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "empresas"))
            )
            dropwdown = Select(seletor_empresa)
            dropwdown.select_by_visible_text(empresa_selecionada)
            time.sleep(2)
            seletor_area = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "areaNegocio"))
            )
            dropwdown = Select(seletor_area)
            dropwdown.select_by_visible_text(area_selecionada)
            time.sleep(2)
            seletor_unidades = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "unidades"))
            )
            dropwdown = Select(seletor_unidades)
            dropwdown.select_by_visible_text(unidade_selecionada)
            time.sleep(2)

            # Pressiona o botão "entrar no SPT"
            botao = driver.find_element(By.CSS_SELECTOR, 'input#sel_un\\:botao')  # Escapando o ":" no seletor
            botao.click()
            print("Botão 'sel_un:botao' pressionado com sucesso.")

            #CLICA EM ANÁLISE DE RISCO E ELABORAR AR
            analise_risco = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//ul[@id='treemenu1']/li/strong[text()='Análise de Risco']")))
            analise_risco.click() # Clica no elemento
            print("Clicou no elemento 'Análise de Risco' com sucesso.")
            elaborar_ar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[@title='Alterar/Excluir/ Visualizar/ Emitir/Revisar']"))
            )
            elaborar_ar.click()
            
            #SELECIONA O STATUS EMITIDA
            try:
                status_ar = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//select[@id='j_id172:status']"))
                )
                dropwdown = Select(status_ar)

                dropwdown.select_by_visible_text('Emitida')
            except Exception as e:
                print(f"Erro ao selecionar uma opção: {e}")
            
            #Inicio da consulta das ARs
            try:
                for index, acao in enumerate(dados_planilha["Ordem"]):  # Itera com índice
                    print(acao)
                    try:
                        if pd.notna(acao):
                            #INSERE O VALOR DA ORDEM DE MANUTENÇÃO
                            campo_codigo_om = WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located((By.XPATH, "//input[@id='j_id172:codigoOrdemManutencao']"))
                            )
                            campo_codigo_om.clear()
                            campo_codigo_om.send_keys(int(dados_planilha["Ordem"][index])) 
                            
                            botao_pesquisar = WebDriverWait(driver, 10).until(
                                EC.element_to_be_clickable((By.XPATH, "//input[@id='j_id172:j_id268']"))
                            )

                            driver.execute_script("arguments[0].scrollIntoView(true);", botao_pesquisar)  # Garante que o botão esteja visível
                            botao_pesquisar.click()

                            WebDriverWait(driver, 10).until(
                                        EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
                                    )
                            
                            #TRATAMENTO DE ERRO PARA AR NÃO ENCONTRADA
                            try:
                                ar_emitida = WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located((By.XPATH, "//div[@id='j_id172:table_tabelaArs:0:j_id677']"))
                                )
                                ar = int(ar_emitida.text)
                                dados_planilha["Número\nA.R."][index] = ar
                                print(f"Número da AR: {ar}")
                                
                            except Exception as e:
                                ar_nao_emitida = WebDriverWait(driver, 10).until(
                                    EC.element_to_be_clickable((By.XPATH, "//div[@id='modalMensagemDiv']/center/input[@value='Fechar']"))
                                )
                                ar_nao_emitida.click()
                                print("Ordem não tem Análise de Risco Emitida")

                                dados_planilha["Número\nA.R."][index] = "ELABORAR AR"
                                continue

                            WebDriverWait(driver, 10).until(
                                        EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
                            )

                        else:
                            continue
                    except Exception as e:
                        dados_planilha["Número\nA.R."][index] = "ELABORAR AR"
                        print("Ordem não está valida para consulta de AR")
                        continue
            except Exception as e:
                print(f"Erro foi: {e}")
        except Exception as e:
            print(f"Erro ao realizar as ações no site: {e}")

        try:
            #SALVAR PANDAS COMO EXCEL
            dados = pd.DataFrame(dados_planilha)
            print(dados)

            # Cria um objeto ExcelWriter
            with pd.ExcelWriter('SemanaAtual.xlsx', engine='openpyxl') as writer:
                dados.to_excel(writer, index=False, sheet_name='Sheet1')

                # Acessa o workbook e a worksheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']

                # Adiciona filtros na primeira linha
                for col in range(len(dados.columns)):
                    worksheet.auto_filter.ref = worksheet.dimensions
                
                # Ajusta a largura das colunas automaticamente
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter  # Obtém a letra da coluna
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

                # Define a quebra de linha automática para todas as células
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

                # Aplica formatação condicional
                green_fill = openpyxl.styles.PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        if cell.column_letter == 'G' and cell.value != 'Não':
                            cell.fill = green_fill
                        if cell.column_letter == 'B' and cell.value == 'Emitir':
                            cell.fill = green_fill
        except Exception as e:
            print(f"Erro ao salvar informações na planilha: {e}")
        
        """#SALVAR INFORMAÇÕES NA PLANILHA
        try:
            #abre o arquivo
            workbook = load_workbook(caminho_arquivo)
            workbook2 = workbook

            #seleciona a planilha
            planilha = workbook2.active

            #percorre as linhas da planilha
            for index, valor in enumerate(dados_planilha["Número\nA.R."], start=2):
                planilha[f"V{index}"] = dados_planilha["Número\nA.R."][index-2]

            #salva o arquivo
            workbook.save(caminho_arquivo)
        except Exception as e:
            print(f"Erro ao salvar informações na planilha: {e}")
            
        #SALVAR PANDAS COMO EXCEL
        dados = pd.DataFrame(dados_planilha)
        print(dados)
        dados.to_excel('SemanaAtual.xlsx', index=False)"""
        
        print("Elaboração finalizada com sucesso!")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
    finally:
        if driver:
            driver.quit()

#ELABORAR AR NA ÁREA
def elaborar_ar_area(caminho_arquivo):
    if not caminho_arquivo:
        print("Nenhum arquivo Excel foi selecionado.")
        return

    dados_planilha = ler_planilha_excel(caminho_arquivo)
    if not dados_planilha:
        return
    
    dados_planilha['AR1_Man'] = ''

#FUNÇÃO PARA ELABORAR AR
def elaborar_ar(usuario, senha, caminho_arquivo):
    if not caminho_arquivo:
        print("Nenhum arquivo Excel foi selecionado.")
        return

    dados_planilha = ler_planilha_excel(caminho_arquivo)
    if not dados_planilha:
        return

    #driver_path = "C:/Users/DMYN/Downloads/edgedriver_win64/msedgedriver.exe"
    #service = Service(driver_path)
    driver = None

    try:
        #driver = webdriver.Edge(service=service)
        driver = webdriver.Firefox(service=Service(GeckoDriverManager().install()))
        driver = realizar_login(driver, usuario, senha)
        acessar_spt(driver)
        clicar_analise_risco(driver)
        elaborar_ars(driver, dados_planilha)
        salvar_informacoes_planilha(dados_planilha, caminho_arquivo)
        #salvar_dados_planilha(dados_planilha, caminho_arquivo)
    except Exception as e:
        print(f"Ocorreu um erro: {e}")
    """finally:
        if driver:
            driver.quit()"""

def clicar_analise_risco(driver):
    try:
        analise_risco = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//ul[@id='treemenu1']/li/strong[text()='Análise de Risco']"))
        )
        analise_risco.click()
        print("Clicou no elemento 'Análise de Risco' com sucesso.")
    except Exception as e:
        print(f"Erro ao clicar em 'Análise de Risco': {e}")

def elaborar_ars(driver, dados_planilha):
    for index, acao in enumerate(dados_planilha["Ordem"]):
        if pd.notna(acao) and "ELABORAR" in dados_planilha["Número\nA.R."][index]:
            try:
                elaborar_ar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Elaborar']"))
                )
                elaborar_ar.click()
                inserir_valor_ordem(driver, dados_planilha, index)
                tratar_ordem_encerrada(driver, dados_planilha, index)
                #preencher_campos_ar(driver, dados_planilha, index)
            except Exception as e:
                print(f"Erro ao elaborar AR para a ordem {acao}: {e}")
                elaborar_ar = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//a[@title='Elaborar']"))
                )
                elaborar_ar.click()
                continue

def inserir_valor_ordem(driver, dados_planilha, index):
    try:
        #XPATH COMPLETO - /html/body/center/div/div[2]/div[2]/form/fieldset/div[1]/table/tbody/tr[4]/td[2]/input
        campo_codigo_om = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "/html/body/center/div/div[2]/div[2]/form/fieldset/div[1]/table/tbody/tr[4]/td[2]/input"))
        )
        campo_codigo_om.clear()
        campo_codigo_om.send_keys(f"00{int(dados_planilha['Ordem'][index])}")
        botao_pesquisar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='j_id179:j_id203']"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", botao_pesquisar)
        botao_pesquisar.click()
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
        )
    except Exception as e:
        print(f"Erro ao inserir valor da ordem: {e}")
        return

def tratar_ordem_encerrada(driver, dados_planilha, index):
    try:
        ordem_encerrada_nplan = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//div[@id='modalMensagemDiv']/center/input[@value='Fechar']"))
        )
        ordem_encerrada_nplan.click()
        print("Ordem não disponível para Emissão de AR")
        return
    except Exception:
        print("Consultando se já está elaborada...")
        tratar_ar_elaborada(driver, dados_planilha, index)
        
def tratar_ar_elaborada(driver, dados_planilha, index):
    try:
        ar_elaborada = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, 
                f"//*[contains(@id, 'j_id172:table_tabelaResultadoOMs:0:j_id198')]"))
        )
        ar_elaborada_text = ar_elaborada.text
        
        if "Associada" in ar_elaborada_text:
            print("AR Emitida")
            # Remover a frase "AR Associada:"
            numero_ar = ar_elaborada_text.replace("AR Associada:", "").strip()
            dados_planilha["Número\nA.R."][index] = numero_ar
            pt_aberta = True
            elaborar_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@title='Elaborar']"))
            )
            elaborar_ar.click()
            return
        print("AR já foi elaborada")
        
    except Exception:
        print("AR disponível para elaboração. Iniciando...")
        preencher_campos_ar(driver, dados_planilha, index)

def preencher_campos_ar(driver, dados_planilha, index):
    try:
        WebDriverWait(driver, 20).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        selecionar_gerencia(driver, dados_planilha, index)
        selecionar_especialidade(driver, dados_planilha, index)
        selecionar_servico(driver, dados_planilha, index)
        selecionar_equipamento(driver, dados_planilha, index)
        preencher_envolvidos(driver, dados_planilha, index)
        clicar_botao_ar1(driver)
        selecionar_opcoes_ar1(driver)
        clicar_botao_ar2(driver)
        copiar_ar_plano(driver, dados_planilha, index)
    except Exception as e:
        print(f"Erro ao preencher campos da AR: {e}")
        elaborar_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@title='Elaborar']"))
        )
        elaborar_ar.click() # Clica no botão "Elaborar"

def selecionar_gerencia(driver, dados_planilha, index):
    try:
        Gerencia = dados_planilha["Gerência Emitente"][index]
        GerenciaEmitente = "G&E/UN-TERM/UTE-NPI/OP" if Gerencia == 'G&E/UN-TERM/UTE-NPI/OP' else "G&E/UN-TERM/UTE-NPI/MAN"
        seletor_gerencia = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "gerenciaCombo"))
        )
        dropwdown = Select(seletor_gerencia)
        dropwdown.select_by_visible_text(GerenciaEmitente)
    except Exception as e:
        print(f"Erro ao selecionar gerência: {e}")

def selecionar_especialidade(driver, dados_planilha, index):
    try:
        especialidade = dados_planilha["ESPECIALIDADES"][index]
        if especialidade in ['ANDAIME', 'CALDEIRARIA']:
            especialidade = "CALDEIRARIA E TUBULAÇÃO"
        elif especialidade in ['AR CONDICIONADO', 'PINTURA']:
            especialidade = "COMPLEMENTAR"
        seletor_especialidade = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "especialidadeCombo"))
        )
        dropwdown = Select(seletor_especialidade)
        dropwdown.select_by_visible_text(especialidade)
    except Exception as e:
        print(f"Erro ao selecionar especialidade: {e}")

def selecionar_servico(driver, dados_planilha, index):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        seletor_servico = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "servicoBasicoCombo"))
        )
        dropwdown = Select(seletor_servico)
        dropwdown.select_by_visible_text('OUTROS')
        campo_outro_servico = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='outroServico']"))
        )
        campo_outro_servico.clear()
        #linhas = dados_planilha["Descrição da PT"][index].split('\n')
        #resultado = [linha.replace('-', '').strip() for linha in linhas if linha.strip()]
        #servico = resultado[1]
        campo_outro_servico.send_keys("Ex: Manutenção Preventiva")
    except Exception as e:
        print(f"Erro ao selecionar serviço: {e}")

def selecionar_equipamento(driver, dados_planilha, index):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        seletor_equipamento = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "tipoEquipamentoCombo"))
        )
        dropwdown = Select(seletor_equipamento)
        dropwdown.select_by_visible_text('OUTROS')
        campo_outro_equipamento = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='outroEquipamento']"))
        )
        campo_outro_equipamento.clear()
        linhas = dados_planilha["Descrição da PT"][index].split('\n')
        resultado = [linha.replace('-', '').strip() for linha in linhas if linha.strip()]
        equipamento = resultado[0]
        campo_outro_equipamento.send_keys(equipamento)
    except Exception as e:
        print(f"Erro ao selecionar equipamento: {e}")

def preencher_envolvidos(driver, dados_planilha, index):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        chaves_man = {
            "ANDAIME": "44583340",
            "AR CONDICIONADO": "TQ24",
            "CALDEIRARIA": "41584160",
            "ELÉTRICA": "TQ25",
            "INSTRUMENTAÇÃO": "X7L1",
            "ISOLAMENTO": "43195216",
            "MECÂNICA": "TQ24",
            "PINTURA": "70119201",
        }
        especialidade = dados_planilha["ESPECIALIDADES"][index]
        print(f"Especialidade: {especialidade}")
        chave = chaves_man.get(especialidade, "TQ24")
        print(f"Chave: {chave}")
        lista_chave = ["tq52", chave, "f7s4"]
        opcoes_participacao = ["Operação*", "Manutenção*", "SMS**"]
        for i, opcao in enumerate(lista_chave):
            campo_chave = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='j_id328:txtChaveOuMatricula']"))
            )
            campo_chave.clear()
            campo_chave.send_keys(opcao)
            tipo_participante = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//select[@id='comboTipoParticipacao']"))
            )
            dropdown4 = Select(tipo_participante)
            opcao_atual = opcoes_participacao[i % len(opcoes_participacao)]
            dropdown4.select_by_visible_text(opcao_atual)
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
            )
            botao_incluir = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[contains(@class, 'botao_incluir') and contains(@onclick, 'abrirPopUpParticipante')]"))
            )
            botao_incluir.click()
            WebDriverWait(driver, 10).until(
                EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))
            )
            time.sleep(1)
    except Exception as e:
        print(f"Erro ao preencher envolvidos: {e}")

def clicar_botao_ar1(driver):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        botao_ar1 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='j_id328:nivel1_man_ope']"))
        )
        botao_ar1.click()
    except Exception as e:
        print(f"Erro ao clicar no botão AR1: {e}")

def selecionar_opcoes_ar1(driver):
    try:
        ar1_opcoes = {
            #OPÇÕES DA MANUTENÇÃO
            "falta procedimento especifico / PBS": "1",
            "Espaço confinado / CO2": "2",
            "Produtos inflamaveis ou toxicos": "3",
            "Instalação eletrica energizada": "4",
            "Alta temperatura ou pressurizado": "5",
            "Mais de uma disciplina": "6",
            "Movimentação Carga critica": "7",
            "escavaçções, perfurações, fundações": "8",
            "Trabalho em altura": "9",
            #OPÇÕES DA OPERAÇÃO
            "impossibilidade de dissipar energia": "10",
            "impossibilidade de aplicar LIBRA": "11",
            "Necessidade de remoção de LIBRA": "12",
        }
        ar1_opcao = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//*[@id='j_id172:tabelaAlterarManutencoesArN1:0:indicadorRespostaM']"))
        )
        dropdown4 = Select(ar1_opcao)
        dropdown4.select_by_visible_text("Sim")
        
        #XPATH BOTAO CONFIRMAR AR1 /html/body/center/div/div[2]/div[2]/form/fieldset/div[3]/input
        confirmar_ar1 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/center/div/div[2]/div[2]/form/fieldset/div[3]/input"))
        )
        confirmar_ar1.click()
        
    except Exception as e:
        print(f"Erro ao selecionar opções AR1: {e}")
        
def clicar_botao_ar2(driver):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        botao_ar2 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='j_id328:nivel2']"))
        )
        botao_ar2.click()
    except Exception as e:
        print(f"Erro ao clicar no botão AR1: {e}")
        
def copiar_ar_plano(driver, dados_planilha, index):
    try:
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        botao_copiar_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//input[contains(@value, 'Copiar Análise de Risco')][2]"))
        )
        botao_copiar_ar.click()
        campo_ar_modelo = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='n_analiseRiscoSelecionado']"))
        )
        campo_ar_modelo.clear()
        campo_ar_modelo.send_keys("5267480") # copiar análise da ordem anterior
        WebDriverWait(driver, 10).until(
            EC.invisibility_of_element((By.ID, "modalPanelIDDiv"))  # ID do elemento bloqueador
        )
        botao_pesquisar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr[2]/td/form/div/div[2]/center/fieldset/table/tbody/tr/td/table/tbody/tr[8]/td/input"))
        )
        botao_pesquisar.click()
        selecao_modelo_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='radios_tabelaAnaliseRiscoNivel2Popup:0']"))
        )
        selecao_modelo_ar.click()
        confirmar_modelo_ar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@id='j_id410:btnConfimarEcluirTarefa']"))
        )
        confirmar_modelo_ar.click()
        retirar_exemplo_tarefa = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//*[@class='btnElaborar3Link']"))
        )
        retirar_exemplo_tarefa.click()
        confirmar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div[2]/div/div[2]/table/tbody/tr/td/div/input[1]"))
        )
        confirmar.click()
        salvar_alteracoes = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/center/div/div[2]/div[2]/form/fieldset/div/div[3]/input[1]"))
        )
        salvar_alteracoes.click()
        texto_ar_salva = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//table[@id='mensagens']/tbody/tr/td/span[starts-with(text(), 'Análise de Risco nº')]"))
        )
        texto_completo = texto_ar_salva.text
        print(f"Texto completo capturado: {texto_completo}")
        match = re.search(r"Análise de Risco nº (\S+\s\S+)", texto_completo)
        if match:
            valor_extraido = match.group(1)
            dados_planilha["Número\nA.R."][index] = valor_extraido
            print(f"O valor da AR que está na planilha é {dados_planilha['Número\nA.R.'][index]}")
            
        else:
            print("Não foi possível extrair o valor do texto.")
        
    except Exception as e:
        print(f"Erro ao clicar no botão AR1: {e}")
        
def salvar_dados_planilha(dados_planilha, caminho_arquivo):
    try:
        dados = pd.DataFrame(dados_planilha)
        dados.to_excel('SemanaAtual.xlsx', index=False)
        workbook = load_workbook(caminho_arquivo)
        planilha = workbook.active
        for index, valor in enumerate(dados_planilha["Número\nA.R."], start=2):
            planilha[f"V{index}"] = dados_planilha["Número\nA.R."][index-2]
        workbook.save(caminho_arquivo)
    except Exception as e:
        print(f"Erro ao salvar informações na planilha: {e}")